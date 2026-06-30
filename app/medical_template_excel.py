"""Загрузка и валидация template_org_medical.xlsx (group_id → log_group)."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from app.medical_org_groups import (
    ALLOWED_EMPTY_GROUP_UNIT_TYPES,
    GROUP_ID_TO_LOG_GROUP,
    group_id_to_log_group,
    normalize_group_id,
)

DEFAULT_MEDICAL_ORG_EXCEL = (
    Path(__file__).resolve().parent / "data" / "universal" / "template_org_medical.xlsx"
)

MAIN_SHEET = "template_org"
GROUPS_SHEET_CANDIDATES = ("Лист1", "Sheet1", "groups")


@dataclass
class MedicalOrgExcelRow:
    code: str
    name: str
    unit_type: str
    parent_code: str | None
    group_id: str | None
    log_group: str | None
    segment_code: str | None = None
    sort_order: int = 0


@dataclass
class MedicalOrgExcelData:
    path: Path
    group_labels: dict[str, str] = field(default_factory=dict)
    rows: list[MedicalOrgExcelRow] = field(default_factory=list)
    by_code: dict[str, MedicalOrgExcelRow] = field(default_factory=dict)
    name_fallback_matches: list[str] = field(default_factory=list)

    def log_group_by_code(self) -> dict[str, str | None]:
        return {code: row.log_group for code, row in self.by_code.items()}


def default_medical_org_excel_path() -> Path:
    return DEFAULT_MEDICAL_ORG_EXCEL


def _header_index(headers: list[str]) -> dict[str, int]:
    return {h.strip(): i for i, h in enumerate(headers) if h and str(h).strip()}


def _parse_groups_from_cell(text: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for m in re.finditer(
        r"(?<![\d|])(\d+)\s*\|\s*([^|\d]+?)(?=\s+\d+\s*\||\(\d+\s*rows?\)|$)",
        text,
        flags=re.IGNORECASE,
    ):
        gid = m.group(1).strip()
        label = m.group(2).strip()
        if gid and label:
            out[gid] = label
    return out


def _load_groups_sheet(wb) -> dict[str, str]:
    ws = None
    for name in wb.sheetnames:
        if name in GROUPS_SHEET_CANDIDATES or name.lower() in ("лист1", "sheet1", "groups"):
            ws = wb[name]
            break
    if ws is None and len(wb.sheetnames) > 1:
        ws = wb[wb.sheetnames[1]]
    if ws is None:
        return _canonical_group_labels()

    groups: dict[str, str] = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _canonical_group_labels()

    first_row = rows[0]
    if len(first_row) >= 2 and first_row[0] is not None and first_row[1] is not None:
        h0 = str(first_row[0]).strip().lower()
        if h0 in ("group_id", "id", "код"):
            for row in rows[1:]:
                if not row or row[0] is None:
                    continue
                gid = normalize_group_id(row[0])
                label = str(row[1] or "").strip()
                if gid and label:
                    groups[gid] = label
            if groups:
                return groups

    blob = "\n".join(str(c) for row in rows for c in row if c is not None)
    parsed = _parse_groups_from_cell(blob)
    if parsed:
        return parsed
    return _canonical_group_labels()


def _canonical_group_labels() -> dict[str, str]:
    from app.medical_org_groups import MEDICAL_LOG_GROUP_LABELS, GROUP_ID_TO_LOG_GROUP

    return {gid: MEDICAL_LOG_GROUP_LABELS[slug] for gid, slug in GROUP_ID_TO_LOG_GROUP.items()}


def validate_medical_org_excel(data: MedicalOrgExcelData) -> list[str]:
    """Ошибки валидации; пустой список — OK."""
    errors: list[str] = []
    known_gids = set(data.group_labels) | set(GROUP_ID_TO_LOG_GROUP)
    seen_codes: set[str] = set()
    for row in data.rows:
        if row.code in seen_codes:
            errors.append(f"duplicate_code:{row.code}")
        seen_codes.add(row.code)
        if row.unit_type in ALLOWED_EMPTY_GROUP_UNIT_TYPES:
            if row.group_id:
                errors.append(f"group_id_on_root:{row.code}")
            continue
        if row.unit_type == "department" and not row.group_id:
            errors.append(f"missing_group_id:department:{row.code}")
        if row.group_id and row.group_id not in known_gids:
            errors.append(f"unknown_group_id:{row.code}:{row.group_id}")
        if row.group_id and row.group_id not in data.group_labels:
            errors.append(f"group_id_not_in_sheet1:{row.code}:{row.group_id}")
    main_gids = {row.group_id for row in data.rows if row.group_id}
    for gid in main_gids:
        if gid not in data.group_labels:
            errors.append(f"group_id_missing_label:{gid}")
    return errors


def load_medical_org_excel(path: Path | None = None) -> MedicalOrgExcelData:
    from openpyxl import load_workbook

    path = path or default_medical_org_excel_path()
    if not path.is_file():
        raise FileNotFoundError(f"medical_org_excel_not_found:{path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = MAIN_SHEET if MAIN_SHEET in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = [
            str(c).strip() if c is not None else ""
            for c in next(ws.iter_rows(max_row=1, values_only=True))
        ]
        idx = _header_index(headers)
        for required in ("code", "unit_type"):
            if required not in idx:
                raise ValueError(f"medical_org_excel_missing_column:{required}")

        gid_col = idx.get("group_id")
        if gid_col is None:
            for k, i in idx.items():
                if k.replace(" ", "") == "group_id":
                    gid_col = i
                    break

        group_labels = _load_groups_sheet(wb)
        data = MedicalOrgExcelData(path=path, group_labels=group_labels)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            code = str(row[idx["code"]] or "").strip()
            if not code:
                continue
            unit_type = str(row[idx["unit_type"]] or "").strip()
            name = str(row[idx.get("name", -1)] or code).strip() if "name" in idx else code
            parent_raw = row[idx["parent_code"]] if "parent_code" in idx else None
            parent_code = str(parent_raw).strip() if parent_raw else None
            seg_raw = row[idx["segment_code"]] if "segment_code" in idx else None
            segment_code = str(seg_raw).strip() if seg_raw else None
            sort_order = int(row[idx["sort_order"]] or 0) if "sort_order" in idx else 0
            group_id = normalize_group_id(row[gid_col]) if gid_col is not None else None
            log_group = None
            if group_id:
                log_group = group_id_to_log_group(group_id)
            elif unit_type not in ALLOWED_EMPTY_GROUP_UNIT_TYPES:
                log_group = None

            item = MedicalOrgExcelRow(
                code=code,
                name=name,
                unit_type=unit_type,
                parent_code=parent_code,
                group_id=group_id,
                log_group=log_group,
                segment_code=segment_code,
                sort_order=sort_order,
            )
            data.rows.append(item)
            data.by_code[code] = item

        errors = validate_medical_org_excel(data)
        if errors:
            raise ValueError("medical_org_excel_validation:" + ";".join(errors))
        return data
    finally:
        wb.close()


def apply_excel_log_groups_to_specs(
    specs: list[dict],
    *,
    path: Path | None = None,
) -> list[dict]:
    """Добавить log_group к spec по коду из Excel (идемпотентно)."""
    try:
        excel = load_medical_org_excel(path)
    except FileNotFoundError:
        return [dict(s) for s in specs]

    out: list[dict] = []
    for spec in specs:
        row = dict(spec)
        code = str(row.get("code") or "").strip()
        ex = excel.by_code.get(code)
        if ex and ex.log_group:
            row["log_group"] = ex.log_group
            row["group_id"] = ex.group_id
        out.append(row)
    return out


def resolve_log_group_for_client_unit(
    *,
    code: str | None,
    catalog_source_code: str | None,
    name: str | None,
    unit_type: str,
    excel: MedicalOrgExcelData,
) -> tuple[str | None, str]:
    """Сопоставить локальный org_unit → log_group. Возвращает (slug, match_kind)."""
    if unit_type in ALLOWED_EMPTY_GROUP_UNIT_TYPES:
        return None, "skip_root"

    for key, kind in (
        (catalog_source_code, "catalog_source_code"),
        (code, "code"),
    ):
        k = str(key or "").strip()
        if not k:
            continue
        ex = excel.by_code.get(k)
        if ex and ex.log_group:
            return ex.log_group, kind

    if name:
        name_norm = str(name).strip().casefold()
        matches = [
            ex
            for ex in excel.rows
            if ex.unit_type == unit_type and ex.name.strip().casefold() == name_norm
        ]
        if len(matches) == 1:
            excel.name_fallback_matches.append(code or name or "?")
            return matches[0].log_group, "name_fallback"
        if len(matches) > 1:
            return None, "name_ambiguous"
    return None, "unmatched"

# route: /api/employees | file: app/routers/employees.py

from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.auth.context import CurrentAccount
from app.auth.deps import get_current_account
from app.auth.tenant import assert_client_access, load_employee_for_ctx, require_client_query_access
from app.db import get_db
from app.excel_export import xlsx_file_response
from app.models import Account, Client, Employee, OrgUnit, Position
from app.schemas import EmployeeCreate, EmployeeListOut, EmployeeOut, EmployeePatch, ListEnvelope
from app.utils import new_id32

router = APIRouter(prefix="/employees", tags=["employees"])


def _assert_org_unit(db: Session, client_id: str, org_unit_id: str | None) -> None:
    if org_unit_id is None:
        return
    ou = db.get(OrgUnit, org_unit_id)
    if not ou or ou.client_id != client_id:
        raise HTTPException(status_code=400, detail="org_unit_not_found")


def _assert_position(db: Session, client_id: str, position_id: str | None) -> None:
    if position_id is None:
        return
    pos = db.get(Position, position_id)
    if not pos or pos.client_id != client_id:
        raise HTTPException(status_code=400, detail="position_not_found")


@router.get("", response_model=ListEnvelope[EmployeeListOut])
def list_employees(
    client_id: str = Query(...),
    org_unit_id: str | None = Query(None),
    position_id: str | None = Query(None),
    search: str | None = Query(None, description="Поиск по ФИО или email"),
    db: Session = Depends(get_db),
    _ctx: CurrentAccount = Depends(require_client_query_access),
    limit: int = Query(100, ge=1, le=2000),
    offset: int = Query(0, ge=0),
) -> ListEnvelope[EmployeeListOut]:
    q = select(Employee).where(Employee.client_id == client_id)
    if org_unit_id:
        q = q.where(Employee.org_unit_id == org_unit_id)
    if position_id:
        q = q.where(Employee.position_id == position_id)
    if search and search.strip():
        term = f"%{search.strip()}%"
        q = q.where(
            or_(
                Employee.last_name.ilike(term),
                Employee.first_name.ilike(term),
                Employee.middle_name.ilike(term),
                Employee.email.ilike(term),
                func.coalesce(Employee.phone, "").ilike(term),
                func.coalesce(Employee.telegram_id, "").ilike(term),
            )
        )
    total = db.scalar(select(func.count()).select_from(q.subquery())) or 0
    rows = db.scalars(q.order_by(Employee.created_at.desc()).limit(limit).offset(offset)).all()
    emp_ids = [r.id for r in rows]
    logins: dict[str, str] = {}
    if emp_ids:
        for acc in db.scalars(select(Account).where(Account.employee_id.in_(emp_ids))).all():
            if acc.employee_id not in logins:
                logins[acc.employee_id] = acc.login
    items = [
        EmployeeListOut(
            **EmployeeOut.model_validate(r).model_dump(),
            account_login=logins.get(r.id),
        )
        for r in rows
    ]
    return ListEnvelope[EmployeeListOut](
        items=items,
        total=total,
        limit=limit,
        offset=offset,
    )


@router.get("/export/excel")
def export_employees_excel(
    client_id: str = Query(...),
    org_unit_id: str | None = Query(None),
    position_id: str | None = Query(None),
    search: str | None = Query(None),
    db: Session = Depends(get_db),
    _ctx: CurrentAccount = Depends(require_client_query_access),
) -> Response:
    if not db.get(Client, client_id):
        raise HTTPException(status_code=404, detail="client_not_found")
    q = select(Employee).where(Employee.client_id == client_id)
    if org_unit_id:
        q = q.where(Employee.org_unit_id == org_unit_id)
    if position_id:
        q = q.where(Employee.position_id == position_id)
    if search and search.strip():
        term = f"%{search.strip()}%"
        q = q.where(
            or_(
                Employee.last_name.ilike(term),
                Employee.first_name.ilike(term),
                Employee.middle_name.ilike(term),
                Employee.email.ilike(term),
                func.coalesce(Employee.phone, "").ilike(term),
                func.coalesce(Employee.telegram_id, "").ilike(term),
            )
        )
    rows = db.scalars(q.order_by(Employee.created_at.desc()).limit(5000)).all()
    emp_ids = [r.id for r in rows]
    logins: dict[str, str] = {}
    if emp_ids:
        for acc in db.scalars(select(Account).where(Account.employee_id.in_(emp_ids))).all():
            if acc.employee_id not in logins:
                logins[acc.employee_id] = acc.login
    headers = [
        "last_name",
        "first_name",
        "middle_name",
        "email",
        "phone",
        "telegram_id",
        "org_unit_code",
        "position_code",
        "employment_status",
        "is_manager",
        "account_login",
        "id",
        "org_unit_id",
        "position_id",
        "client_id",
        "created_at",
        "updated_at",
    ]
    data = []
    for r in rows:
        ou = db.get(OrgUnit, r.org_unit_id) if r.org_unit_id else None
        pos = db.get(Position, r.position_id) if r.position_id else None
        data.append(
            [
                r.last_name,
                r.first_name,
                r.middle_name,
                r.email,
                r.phone,
                r.telegram_id,
                ou.code if ou else None,
                pos.code if pos else None,
                r.employment_status,
                r.is_manager,
                logins.get(r.id),
                r.id,
                r.org_unit_id,
                r.position_id,
                r.client_id,
                r.created_at,
                r.updated_at,
            ]
        )
    return xlsx_file_response(
        download_name=f"employees_{client_id}.xlsx",
        sheet_title="employees",
        headers=headers,
        rows=data,
    )


@router.get("/{employee_id}", response_model=EmployeeOut)
def get_employee(
    employee_id: str,
    db: Session = Depends(get_db),
    ctx: CurrentAccount = Depends(get_current_account),
) -> EmployeeOut:
    obj = load_employee_for_ctx(db, employee_id, ctx)
    return EmployeeOut.model_validate(obj)


@router.post("", response_model=EmployeeOut)
def create_employee(
    payload: EmployeeCreate,
    db: Session = Depends(get_db),
    ctx: CurrentAccount = Depends(get_current_account),
) -> EmployeeOut:
    assert_client_access(ctx, payload.client_id)
    _assert_org_unit(db, payload.client_id, payload.org_unit_id)
    _assert_position(db, payload.client_id, payload.position_id)
    obj = Employee(
        id=payload.id or new_id32(),
        client_id=payload.client_id,
        last_name=payload.last_name,
        first_name=payload.first_name,
        middle_name=payload.middle_name,
        email=payload.email,
        phone=payload.phone,
        telegram_id=payload.telegram_id,
        org_unit_id=payload.org_unit_id,
        position_id=payload.position_id,
        employment_status=payload.employment_status,
        is_manager=payload.is_manager,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return EmployeeOut.model_validate(obj)


@router.patch("/{employee_id}", response_model=EmployeeOut)
def patch_employee(
    employee_id: str,
    payload: EmployeePatch,
    db: Session = Depends(get_db),
    ctx: CurrentAccount = Depends(get_current_account),
) -> EmployeeOut:
    obj = load_employee_for_ctx(db, employee_id, ctx)
    data = payload.model_dump(exclude_unset=True)
    if "org_unit_id" in data:
        _assert_org_unit(db, obj.client_id, data["org_unit_id"])
    if "position_id" in data:
        _assert_position(db, obj.client_id, data["position_id"])
    for k, v in data.items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return EmployeeOut.model_validate(obj)


@router.delete("/{employee_id}", status_code=204)
def delete_employee(
    employee_id: str,
    db: Session = Depends(get_db),
    ctx: CurrentAccount = Depends(get_current_account),
) -> Response:
    obj = load_employee_for_ctx(db, employee_id, ctx)
    from app.models import Account
    acc = db.scalar(select(Account).where(Account.employee_id == employee_id))
    if acc:
        raise HTTPException(status_code=400, detail="employee_has_account")
    db.delete(obj)
    db.commit()
    return Response(status_code=204)


@router.post("/import-excel", response_model=list[EmployeeOut])
def import_employees_excel(
    client_id: str = Query(...),
    sheet: int | str | None = Query(None, description="Номер листа (1, 2, 3...) или имя. По умолчанию — первый лист."),
    header_row: int = Query(1, ge=1, description="Строка с заголовками (1, 2, 3...). По умолчанию — 1."),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _ctx: CurrentAccount = Depends(require_client_query_access),
) -> list[EmployeeOut]:
    """Импорт сотрудников из Excel. Колонки: last_name, first_name, middle_name, email (или Фамилия, Имя, Отчество, Email)."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="expected_xlsx_or_xls")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=503, detail="install_openpyxl_for_excel_import")
    content = file.file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"invalid_excel: {e!s}") from e
    if sheet is not None:
        if isinstance(sheet, int):
            if sheet < 1 or sheet > len(wb.worksheets):
                raise HTTPException(
                    status_code=400,
                    detail=f"sheet_index_out_of_range: available 1..{len(wb.worksheets)}",
                )
            ws = wb.worksheets[sheet - 1]
        else:
            if sheet not in wb.sheetnames:
                raise HTTPException(
                    status_code=400,
                    detail=f"sheet_not_found: {sheet!r}. Available: {wb.sheetnames}",
                )
            ws = wb[sheet]
    else:
        ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="empty_excel")

    def _build_col_map(rows: list, hdr_idx: int) -> tuple[dict, list[str], bool, bool]:
        if hdr_idx >= len(rows):
            return {}, [], False, False
        headers_raw = [str(h).strip() if h else "" for h in rows[hdr_idx]]
        headers = [h.lower() for h in headers_raw]
        headers_norm = [h.replace(" ", "").replace(".", "").replace("-", "") for h in headers]
        col_map = {}
        for name, alt in [
            ("id", ["id", "employee_id", "employeeid", "ид", "идентификатор", "кодсотрудника"]),
            ("last_name", ["last_name", "lastname", "last name", "фамилия", "surname", "familia"]),
            ("first_name", ["first_name", "firstname", "first name", "имя", "name", "imya"]),
            ("middle_name", ["middle_name", "middlename", "middle name", "отчество", "patronymic", "otchestvo"]),
            ("email", ["email", "почта", "e-mail", "mail", "эл.почта", "электронная почта"]),
            ("phone", ["phone", "телефон", "mobile", "мобильный", "сотовый", "мобтел"]),
            ("telegram_id", ["telegram_id", "telegram", "телеграм", "tg", "telegram id"]),
            ("org_unit_id", ["org_unit_id", "orgunitid", "department_id", "departmentid", "idподразделения", "подразделение_id"]),
            ("org_unit_code", ["org_unit_code", "кодподразделения", "код_подразделения", "orgunitcode", "department_code", "departmentcode"]),
            ("org_unit_name", ["org_unit_name", "подразделение", "отделение", "отдел", "служба", "департамент", "department", "org_unit", "org unit", "orgunit"]),
            ("position_id", ["position_id", "positionid", "job_id", "idдолжности", "должность_id"]),
            ("position_code", ["position_code", "коддолжности", "код_должности", "positioncode"]),
            ("position_name", ["position_name", "должность", "позиция", "position", "job_title", "job title", "jobtitle"]),
            ("fio", ["сотрудник", "фИО", "ф.и.о.", "fio", "employee", "full name", "fullname", "фио"]),
        ]:
            for a in alt:
                anorm = a.replace(" ", "").replace(".", "").replace("-", "")
                for i, (h, hn) in enumerate(zip(headers, headers_norm)):
                    if a == h or anorm == hn:
                        col_map[name] = i
                        break
                if name in col_map:
                    break
        has_sep = "last_name" in col_map and "first_name" in col_map
        has_f = "fio" in col_map
        return col_map, headers_raw, has_sep, has_f

    sheets_to_try: list[tuple[object, str]] = []
    if sheet is not None:
        sheets_to_try = [(ws, ws.title if hasattr(ws, "title") else "?")]
    else:
        for s in wb.worksheets:
            sheets_to_try.append((s, s.title))

    ws, ws_title, rows, header_idx, col_map, has_separate, has_fio = None, "", [], 0, {}, False, False
    for candidate_ws, candidate_title in sheets_to_try:
        cand_rows = list(candidate_ws.iter_rows(values_only=True))
        if not cand_rows:
            continue
        hdr_idx = header_row - 1
        if hdr_idx >= len(cand_rows):
            continue
        cm, found, hs, hf = _build_col_map(cand_rows, hdr_idx)
        if hs or hf:
            ws, ws_title, rows, header_idx, col_map, has_separate, has_fio = (
                candidate_ws, candidate_title, cand_rows, hdr_idx, cm, hs, hf
            )
            break

    if not ws or not rows:
        raise HTTPException(status_code=400, detail="empty_excel")
    if not has_separate and not has_fio:
        first_ws = sheets_to_try[0][0]
        first_rows = list(first_ws.iter_rows(values_only=True))
        hdr_idx = header_row - 1
        all_found = [str(h).strip() for h in first_rows[hdr_idx] if h] if hdr_idx < len(first_rows) else []
        hint = ""
        if any("должность" in (h or "").lower() for h in all_found):
            hint = " Первый лист — «Должности». Укажите имя листа со списком сотрудников (с колонкой «Сотрудник») в поле «Лист», или переставьте листы в Excel: лист со сотрудниками — первым."
        raise HTTPException(
            status_code=400,
            detail=f"required_columns: last_name+first_name или Сотрудник (ФИО). Найдены: {all_found!r}.{hint}",
        )

    def _parse_fio(val: str) -> tuple[str | None, str | None, str | None]:
        """Парсит ФИО: 'Фамилия Имя Отчество' или 'Фамилия Имя Отчество (осн.)'."""
        if not val or not str(val).strip():
            return None, None, None
        s = str(val).strip()
        for suffix in ("(осн.)", "(осн)", "(основной)", "(основная)"):
            if suffix in s:
                s = s.replace(suffix, "").strip()
        parts = s.split()
        if len(parts) >= 2:
            return parts[0], parts[1], " ".join(parts[2:]) if len(parts) > 2 else None
        return None, None, None

    def _norm_lookup_key(value: object) -> str:
        return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())

    def _cell_str(row: tuple, key: str) -> str:
        idx = col_map.get(key)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    org_units = db.scalars(select(OrgUnit).where(OrgUnit.client_id == client_id)).all()
    positions = db.scalars(select(Position).where(Position.client_id == client_id)).all()
    org_units_by_id = {x.id: x for x in org_units}
    org_units_by_code = {_norm_lookup_key(x.code): x for x in org_units if x.code}
    org_units_by_name = {_norm_lookup_key(x.name): x for x in org_units if x.name}
    positions_by_id = {x.id: x for x in positions}
    positions_by_code_by_org: dict[tuple[str, str], Position] = {}
    positions_by_name_by_org: dict[tuple[str, str], Position] = {}
    positions_by_code_all: dict[str, list[Position]] = {}
    positions_by_name_all: dict[str, list[Position]] = {}
    for p in positions:
        for val in (p.code, getattr(p, "position_catalog_code", None)):
            key = _norm_lookup_key(val)
            if key:
                positions_by_code_by_org.setdefault((key, p.org_unit_id), p)
                positions_by_code_all.setdefault(key, []).append(p)
        key = _norm_lookup_key(p.name)
        if key:
            positions_by_name_by_org.setdefault((key, p.org_unit_id), p)
            positions_by_name_all.setdefault(key, []).append(p)
    positions_by_unique_code = {k: v[0] for k, v in positions_by_code_all.items() if len(v) == 1}
    positions_by_unique_name = {k: v[0] for k, v in positions_by_name_all.items() if len(v) == 1}

    def _resolve_org_unit(row: tuple) -> OrgUnit | None:
        raw_code = _cell_str(row, "org_unit_code")
        raw_name = _cell_str(row, "org_unit_name")
        for raw in (raw_code, raw_name):
            key = _norm_lookup_key(raw)
            if not key:
                continue
            found = org_units_by_code.get(key) or org_units_by_name.get(key)
            if found:
                return found
        raw_id = _cell_str(row, "org_unit_id")
        if raw_id:
            return org_units_by_id.get(raw_id)
        return None

    def _resolve_position(row: tuple, org_unit_id: str | None) -> Position | None:
        raw_code = _cell_str(row, "position_code")
        raw_name = _cell_str(row, "position_name")
        if org_unit_id:
            for raw, by_org in ((raw_code, positions_by_code_by_org), (raw_name, positions_by_name_by_org)):
                key = _norm_lookup_key(raw)
                if not key:
                    continue
                found = by_org.get((key, org_unit_id))
                if found:
                    return found
            if raw_code or raw_name:
                return None
        else:
            for raw, by_unique in ((raw_code, positions_by_unique_code), (raw_name, positions_by_unique_name)):
                key = _norm_lookup_key(raw)
                if not key:
                    continue
                found = by_unique.get(key)
                if found:
                    return found
        raw_id = _cell_str(row, "position_id")
        if raw_id:
            found = positions_by_id.get(raw_id)
            if found and (org_unit_id is None or found.org_unit_id == org_unit_id):
                return found
        return None

    def _find_existing_employee(
        employee_id: str | None,
        last_name: str,
        first_name: str,
        middle_name: str | None,
        email: str | None,
    ) -> Employee | None:
        if employee_id:
            by_id = db.get(Employee, employee_id)
            if by_id and by_id.client_id == client_id:
                return by_id
        if email:
            by_email = db.scalar(
                select(Employee).where(Employee.client_id == client_id, func.lower(Employee.email) == email.lower())
            )
            if by_email:
                return by_email
        return db.scalar(
            select(Employee)
            .where(
                Employee.client_id == client_id,
                func.lower(Employee.last_name) == last_name.lower(),
                func.lower(Employee.first_name) == first_name.lower(),
                func.coalesce(func.lower(Employee.middle_name), "") == (middle_name or "").lower(),
            )
            .order_by(Employee.created_at.desc())
        )

    created: list[EmployeeOut] = []
    data_rows = rows[header_idx + 1 :]
    for row in data_rows:
        if not row:
            continue
        if has_fio and not has_separate:
            fio_val = row[col_map["fio"]] if col_map["fio"] < len(row) else None
            last_name, first_name, middle_name = _parse_fio(fio_val)
        else:
            last_name = row[col_map["last_name"]] if col_map["last_name"] < len(row) else None
            first_name = row[col_map["first_name"]] if col_map["first_name"] < len(row) else None
            last_name = str(last_name).strip() if last_name else ""
            first_name = str(first_name).strip() if first_name else ""
            middle_name = None
            if "middle_name" in col_map and col_map["middle_name"] is not None and col_map["middle_name"] < len(row) and row[col_map["middle_name"]]:
                middle_name = str(row[col_map["middle_name"]]).strip() or None
        if not last_name or not first_name:
            continue
        email = None
        if "email" in col_map and col_map["email"] is not None and col_map["email"] < len(row) and row[col_map["email"]]:
            email = str(row[col_map["email"]]).strip() or None
        phone = None
        if "phone" in col_map and col_map["phone"] is not None and col_map["phone"] < len(row) and row[col_map["phone"]]:
            phone = str(row[col_map["phone"]]).strip() or None
        telegram_id = None
        if "telegram_id" in col_map and col_map["telegram_id"] is not None and col_map["telegram_id"] < len(row) and row[col_map["telegram_id"]]:
            telegram_id = str(row[col_map["telegram_id"]]).strip() or None
        employee_id = _cell_str(row, "id") or None
        org_unit_id = None
        ou = _resolve_org_unit(row)
        if ou:
            org_unit_id = ou.id
        position_id = None
        pos = _resolve_position(row, org_unit_id)
        if pos:
            position_id = pos.id
            org_unit_id = pos.org_unit_id
        obj = _find_existing_employee(employee_id, last_name, first_name, middle_name, email)
        if obj:
            obj.last_name = last_name
            obj.first_name = first_name
            obj.middle_name = middle_name
            if email:
                obj.email = email
            if phone:
                obj.phone = phone
            if telegram_id:
                obj.telegram_id = telegram_id
            if org_unit_id:
                obj.org_unit_id = org_unit_id
            if position_id:
                obj.position_id = position_id
        else:
            obj = Employee(
                id=new_id32(),
                client_id=client_id,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                email=email,
                phone=phone,
                telegram_id=telegram_id,
                org_unit_id=org_unit_id,
                position_id=position_id,
                employment_status="active",
                is_manager=False,
            )
            db.add(obj)
        db.flush()
        created.append(EmployeeOut.model_validate(obj))
    db.commit()
    return created


@router.post("/bulk", response_model=list[EmployeeOut])
def bulk_upsert_employees(
    items: list[EmployeeCreate],
    db: Session = Depends(get_db),
    ctx: CurrentAccount = Depends(get_current_account),
) -> list[EmployeeOut]:
    out: list[EmployeeOut] = []
    for it in items:
        assert_client_access(ctx, it.client_id)
        obj = db.get(Employee, it.id) if it.id else None
        _assert_org_unit(db, it.client_id, it.org_unit_id)
        _assert_position(db, it.client_id, it.position_id)
        if obj:
            if obj.client_id != it.client_id:
                raise HTTPException(status_code=400, detail="client_mismatch")
            obj.last_name = it.last_name
            obj.first_name = it.first_name
            obj.middle_name = it.middle_name
            obj.email = it.email
            obj.phone = it.phone
            obj.telegram_id = it.telegram_id
            obj.org_unit_id = it.org_unit_id
            obj.position_id = it.position_id
            obj.employment_status = it.employment_status
            obj.is_manager = it.is_manager
        else:
            obj = Employee(
                id=it.id or new_id32(),
                client_id=it.client_id,
                last_name=it.last_name,
                first_name=it.first_name,
                middle_name=it.middle_name,
                email=it.email,
                phone=it.phone,
                telegram_id=it.telegram_id,
                org_unit_id=it.org_unit_id,
                position_id=it.position_id,
                employment_status=it.employment_status,
                is_manager=it.is_manager,
            )
            db.add(obj)
        db.flush()
        out.append(EmployeeOut.model_validate(obj))
    db.commit()
    return out


"""Загрузка универсальных KPI из XLSX и дополнений к регламентам из JSON (для Docker / первого запуска)."""

from __future__ import annotations

import json
from pathlib import Path

from uuid import NAMESPACE_URL, uuid5

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import KpiTemplate, PositionRegulation, RegulationKpi

# Пути относительно корня репозитория и образа (/app)
_APP_DIR = Path(__file__).resolve().parent
_REPO_ROOT = _APP_DIR.parent


def _kpi_xlsx_paths() -> list[Path]:
    paths = [
        _APP_DIR / "data" / "universal" / "kpi_templates1.xlsx",
        _REPO_ROOT / "docs" / "regulations" / "kpi_templates1.xlsx",
    ]
    return [p for p in paths if p.is_file()]


def _regulation_json_path() -> Path | None:
    p = _APP_DIR / "data" / "universal" / "regulations_enrichment.json"
    return p if p.is_file() else None


def _client_regulations_xlsx_paths() -> list[Path]:
    """Экспорт клиентских регламентов как источник текстов по умолчанию для глобального справочника."""
    paths = [
        _APP_DIR / "data" / "universal" / "client_regulations.xlsx",
        _REPO_ROOT / "docs" / "regulations" / "client_regulations.xlsx",
    ]
    return [p for p in paths if p.is_file()]


def merge_kpi_templates_from_xlsx(db: Session) -> int:
    """
    Вставить из XLSX строки, которых ещё нет в kpi_templates (по kpi_code).
    Ожидаемый лист: kpi_templates; колонки как в шаблоне импорта.
    """
    paths = _kpi_xlsx_paths()
    if not paths:
        return 0
    path = paths[0]
    wb = load_workbook(path, read_only=True, data_only=True)
    if "kpi_templates" not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb["kpi_templates"]
    rows = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return 0
    idx = {str(c).strip(): i for i, c in enumerate(header) if c is not None}
    need = "kpi_code"
    if need not in idx:
        wb.close()
        return 0

    def col(name: str, default: int | None = None) -> int | None:
        return idx.get(name, default)

    i_code = col("kpi_code")
    i_name = col("kpi_name")
    i_unit = col("unit")
    i_period = col("period_type")
    i_formula = col("formula_or_rule")
    i_target = col("default_target")
    i_active = col("is_active")
    i_pos = col("position_code")

    existing = set(db.scalars(select(KpiTemplate.kpi_code)).all())
    created = 0
    for row in rows:
        if not row or i_code is None:
            continue
        code = row[i_code]
        if code is None or str(code).strip() == "":
            continue
        kpi_code = str(code).strip()
        if kpi_code in existing:
            continue
        name = str(row[i_name]).strip() if i_name is not None and row[i_name] is not None else kpi_code
        unit = "%"
        if i_unit is not None and row[i_unit] is not None:
            unit = str(row[i_unit]).strip() or "%"
        period_type = "month"
        if i_period is not None and row[i_period] is not None:
            period_type = str(row[i_period]).strip() or "month"
        formula = None
        if i_formula is not None and row[i_formula] is not None:
            formula = str(row[i_formula]).strip() or None
        default_target = None
        if i_target is not None and row[i_target] is not None:
            try:
                default_target = float(row[i_target])
            except (TypeError, ValueError):
                default_target = None
        is_active = True
        if i_active is not None and row[i_active] is not None:
            v = row[i_active]
            if isinstance(v, bool):
                is_active = v
            else:
                is_active = str(v).strip().lower() in ("1", "true", "yes", "да", "y")
        position_code = None
        if i_pos is not None and row[i_pos] is not None:
            pc = str(row[i_pos]).strip()
            position_code = pc or None

        db.add(
            KpiTemplate(
                kpi_code=kpi_code,
                kpi_name=name[:256],
                unit=unit[:32],
                period_type=period_type[:16],
                formula_or_rule=(formula[:512] if formula else None),
                default_target=default_target,
                is_active=is_active,
                position_code=position_code[:64] if position_code else None,
            )
        )
        existing.add(kpi_code)
        created += 1
    wb.close()
    if created:
        db.commit()
    return created


def apply_regulation_enrichment_json(db: Session) -> int:
    """
    Обновить существующие глобальные регламенты полями из JSON (цели, ЦКП, ссылки).
    Формат: {"regulations": [{"regulation_code": "...", "goal_summary": "...", ...}, ...]}
    """
    path = _regulation_json_path()
    if not path:
        return 0
    data = json.loads(path.read_text(encoding="utf-8"))
    items = data.get("regulations") or []
    updated = 0
    for item in items:
        code = (item.get("regulation_code") or "").strip()
        if not code:
            continue
        obj = db.scalar(select(PositionRegulation).where(PositionRegulation.regulation_code == code))
        if not obj:
            continue
        for field in (
            "regulation_name",
            "goal_summary",
            "ckp_short",
            "ckp_full",
            "google_doc_url",
            "instructions_folder_url",
        ):
            if field in item and item[field] is not None:
                val = item[field]
                s = val if isinstance(val, str) else str(val)
                if field == "regulation_name":
                    s = s[:256]
                elif field in ("goal_summary", "ckp_short"):
                    s = s[:512]
                elif field in ("google_doc_url", "instructions_folder_url"):
                    s = s[:512]
                setattr(obj, field, s)
        updated += 1
    if updated:
        db.commit()
    return updated


def merge_position_regulations_from_client_xlsx(db: Session) -> int:
    """
    Применить к глобальным position_regulations поля из выгрузки client_regulations.xlsx
    (по regulation_code / global_regulation_code). Имеет приоритет над JSON-обогащением
    для указанных полей. Ожидается лист client_regulations; строки с разными client_id
    с одинаковым кодом регламента обычно дублируют текст — берётся первая встреченная строка.
    """
    paths = _client_regulations_xlsx_paths()
    if not paths:
        return 0
    path = paths[0]
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "client_regulations" if "client_regulations" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return 0
    idx = {str(c).strip(): i for i, c in enumerate(header) if c is not None}

    def col(name: str) -> int | None:
        return idx.get(name)

    i_reg = col("regulation_code")
    i_glob = col("global_regulation_code")
    if i_reg is None and i_glob is None:
        wb.close()
        return 0

    fields_text = ("goal_summary", "ckp_short", "ckp_full", "google_doc_url", "instructions_folder_url", "regulation_name")
    col_idx = {f: col(f) for f in fields_text}

    seen_codes: set[str] = set()
    updated = 0

    def norm_cell(row: tuple, i: int | None) -> str | None:
        if i is None or not row:
            return None
        if i >= len(row):
            return None
        v = row[i]
        if v is None:
            return None
        if isinstance(v, float) and v == int(v):
            v = int(v)
        s = str(v).strip()
        return s if s else None

    for row in rows:
        if not row:
            continue
        code = norm_cell(row, i_reg) or norm_cell(row, i_glob)
        if not code:
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)

        obj = db.scalar(select(PositionRegulation).where(PositionRegulation.regulation_code == code))
        if not obj:
            continue

        if col_idx["regulation_name"] is not None:
            s = norm_cell(row, col_idx["regulation_name"])
            if s is not None:
                obj.regulation_name = s[:256]
        if col_idx["goal_summary"] is not None:
            s = norm_cell(row, col_idx["goal_summary"])
            if s is not None:
                obj.goal_summary = s[:512]
        if col_idx["ckp_short"] is not None:
            s = norm_cell(row, col_idx["ckp_short"])
            if s is not None:
                obj.ckp_short = s[:512]
        if col_idx["ckp_full"] is not None:
            s = norm_cell(row, col_idx["ckp_full"])
            if s is not None:
                obj.ckp_full = s
        if col_idx["google_doc_url"] is not None:
            s = norm_cell(row, col_idx["google_doc_url"])
            if s is not None:
                obj.google_doc_url = s[:512]
        if col_idx["instructions_folder_url"] is not None:
            s = norm_cell(row, col_idx["instructions_folder_url"])
            if s is not None:
                obj.instructions_folder_url = s[:512]
        updated += 1

    wb.close()
    if updated:
        db.commit()
    return updated


def _reg_kpi_id(regulation_code: str, kpi_code: str) -> str:
    return uuid5(NAMESPACE_URL, f"seed:reg_kpi:{regulation_code}:{kpi_code}").hex


def link_regulation_kpis_from_templates(db: Session) -> int:
    """
    Для каждого глобального регламента связать шаблоны KPI с тем же position_code.
    Плейсхолдеры KPI_TMPL_* используются только если других шаблонов с этой должностью нет.
    """
    existing = {(r.regulation_code, r.kpi_code) for r in db.scalars(select(RegulationKpi)).all()}
    by_pos: dict[str, list[KpiTemplate]] = {}
    for t in db.scalars(select(KpiTemplate)).all():
        pc = (t.position_code or "").strip()
        if not pc:
            continue
        by_pos.setdefault(pc, []).append(t)

    created = 0
    for reg in db.scalars(select(PositionRegulation)).all():
        templates = by_pos.get(reg.position_code, [])
        if not templates:
            continue
        chosen = [t for t in templates if not t.kpi_code.startswith("KPI_TMPL_")]
        use = chosen if chosen else templates
        for t in use:
            key = (reg.regulation_code, t.kpi_code)
            if key in existing:
                continue
            db.add(
                RegulationKpi(
                    id=_reg_kpi_id(reg.regulation_code, t.kpi_code),
                    regulation_code=reg.regulation_code,
                    kpi_code=t.kpi_code,
                    target_value=t.default_target,
                    period_type=t.period_type,
                    weight=None,
                    is_required=True,
                )
            )
            existing.add(key)
            created += 1
    if created:
        db.commit()
    return created

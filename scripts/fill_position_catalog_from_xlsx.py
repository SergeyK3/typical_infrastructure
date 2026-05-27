#!/usr/bin/env python3
"""Заполнить position_name_en и default_regulation_code в position_catalog из Excel."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.db import SessionLocal
from app.models import PositionCatalog, PositionRegulation
from app.position_name_en import HOSP_POSITION_NAME_EN, position_name_en_for

# Должности шаблона hosp, которых нет в default xlsx (алиас для обратной совместимости)
HOSP_NAME_EN = HOSP_POSITION_NAME_EN


def load_xlsx_rows(path: Path, template_code: str) -> dict[str, dict[str, str | None]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers)}
    need = {"position_code", "position_name_en", "default_regulation_code"}
    if not need.issubset(idx):
        raise ValueError(f"Missing columns in {path}: {need - set(idx)}")

    out: dict[str, dict[str, str | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        tpl = (row[idx.get("template_code", -1)] or template_code)
        if str(tpl).strip() != template_code:
            continue
        code = (row[idx["position_code"]] or "")
        code = str(code).strip()
        if not code:
            continue
        en = row[idx["position_name_en"]]
        reg = row[idx["default_regulation_code"]]
        out[code] = {
            "position_name_en": str(en).strip() if en else None,
            "default_regulation_code": str(reg).strip() if reg else None,
        }
    wb.close()
    return out


def current_regulation_by_position(db, template_code: str) -> dict[str, str]:
    """Актуальный регламент по должности из position_regulations."""
    mapping: dict[str, str] = {}
    for reg in db.scalars(
        select(PositionRegulation).where(
            PositionRegulation.template_code == template_code,
            PositionRegulation.is_current == True,
        )
    ).all():
        code = (reg.position_code or "").strip()
        rcode = (reg.regulation_code or "").strip()
        if code and rcode and code not in mapping:
            mapping[code] = rcode
    return mapping


def fill_template(
    db,
    template_code: str,
    xlsx_by_code: dict[str, dict[str, str | None]],
    *,
    dry_run: bool,
) -> tuple[int, int]:
    regs = current_regulation_by_position(db, template_code)
    updated = 0
    skipped = 0
    rows = db.scalars(
        select(PositionCatalog).where(PositionCatalog.template_code == template_code)
    ).all()
    for row in rows:
        code = row.position_code.strip()
        src = xlsx_by_code.get(code, {})
        new_en = (src.get("position_name_en") or "").strip() or position_name_en_for(template_code, code)
        new_reg = (src.get("default_regulation_code") or "").strip() or regs.get(code)

        cur_en = (row.position_name_en or "").strip()
        cur_reg = (row.default_regulation_code or "").strip()
        if not new_en and not new_reg:
            skipped += 1
            continue
        changed = False
        if new_en and new_en != cur_en:
            row.position_name_en = new_en
            changed = True
        if new_reg and new_reg != cur_reg:
            row.default_regulation_code = new_reg
            changed = True
        if changed:
            updated += 1
            print(f"  {code}: en={row.position_name_en!r} reg={row.default_regulation_code!r}")
        else:
            skipped += 1
    if not dry_run and updated:
        db.flush()
    return updated, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "xlsx",
        nargs="?",
        default=str(Path.home() / "Downloads" / "position_catalog_default_filled.xlsx"),
        help="Excel с заполненным каталогом default",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    path = Path(args.xlsx)
    if not path.is_file():
        raise SystemExit(f"File not found: {path}")

    xlsx_default = load_xlsx_rows(path, "default")
    print(f"Loaded {len(xlsx_default)} rows from {path} (template_code=default)")

    db = SessionLocal()
    try:
        for tpl in ("default", "hosp"):
            print(f"\n=== {tpl} ===")
            n, skip = fill_template(db, tpl, xlsx_default, dry_run=args.dry_run)
            print(f"updated={n} skipped={skip}")
        if not args.dry_run:
            db.commit()
            print("\nCommitted.")
        else:
            print("\nDry run — no commit.")
    finally:
        db.close()


if __name__ == "__main__":
    main()

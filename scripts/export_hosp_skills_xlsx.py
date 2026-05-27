#!/usr/bin/env python3
"""
Excel-реестр навыков шаблона стационара (hosp): 3 столбца из DOCX регламентов.

  - Название должности
  - Твердый или мягкий навык
  - Название навыка

Источники:
  - docs/regulations/gdrive_hosp  — медицинский контур ММЦ (15 должностей)
  - docs/regulations/gdrive_default — административные / back-office регламенты

Выход: docs/regulations/hosp_skills.xlsx

Запуск из корня репозитория:
  python scripts/export_hosp_skills_xlsx.py
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.db import SessionLocal
from app.models import PositionCatalog
from scripts.fix_hosp_regulation_codes_and_urls import HOSP_REG_CANON
from scripts.sync_global_regulations_from_sources import (
    DEFAULT_POSITION_MAP,
    ParsedRegulation,
    _target_position,
    collect_parsed,
    load_url_maps,
)

OUT_PATH = ROOT / "docs" / "regulations" / "hosp_skills.xlsx"
HEADERS = ("Название должности", "Твердый или мягкий навык", "Название навыка")

MMC_POSITION_CODES = frozenset(spec["position_code"] for spec in HOSP_REG_CANON)

# Порядок функциональных блоков для административного контура в xlsx
ADMIN_FUNCTION_ORDER = (
    "ADM",
    "HR",
    "ACC",
    "MKT",
    "LEAD",
    "SALES",
    "QUAL",
    "PR",
    "INFO",
    "PROD",
)


def _parsed_for_position(by_pos: dict[str, ParsedRegulation], catalog_pos: str) -> ParsedRegulation | None:
    if catalog_pos in by_pos:
        return by_pos[catalog_pos]
    for docx_pos, mapped in DEFAULT_POSITION_MAP.items():
        if mapped == catalog_pos and docx_pos in by_pos:
            return by_pos[docx_pos]
    for docx_pos, row in by_pos.items():
        if _target_position("hosp", docx_pos) == catalog_pos:
            return row
    return None


def _append_skills(
    out: list[tuple[str, str, str]],
    *,
    pos_name: str,
    parsed: ParsedRegulation,
) -> None:
    for _rank, title in parsed.hard_skills:
        out.append((pos_name, "твердый", title.strip()))
    for _rank, title in parsed.soft_skills:
        out.append((pos_name, "мягкий", title.strip()))


def collect_rows() -> list[tuple[str, str, str]]:
    by_pos = collect_parsed(load_url_maps())
    db = SessionLocal()
    out: list[tuple[str, str, str]] = []
    seen_positions: set[str] = set()
    try:
        # 1. Медицинский контур ММЦ (gdrive_hosp)
        for spec in HOSP_REG_CANON:
            pos_code = spec["position_code"]
            pc = db.get(PositionCatalog, ("hosp", pos_code))
            pos_name = (pc.position_name_ru if pc else pos_code).strip()
            parsed = _parsed_for_position(by_pos, pos_code)
            if not parsed or not (parsed.hard_skills or parsed.soft_skills):
                continue
            _append_skills(out, pos_name=pos_name, parsed=parsed)
            seen_positions.add(pos_code)

        # 2. Административные регламенты (gdrive_default) для должностей hosp вне ММЦ
        admin_positions = db.scalars(
            select(PositionCatalog)
            .where(PositionCatalog.template_code == "hosp", PositionCatalog.is_active == True)
            .order_by(PositionCatalog.position_code)
        ).all()
        admin_positions = [
            pc
            for pc in admin_positions
            if pc.position_code not in MMC_POSITION_CODES and pc.position_code not in seen_positions
        ]
        admin_positions.sort(
            key=lambda pc: (
                ADMIN_FUNCTION_ORDER.index(pc.function_code)
                if pc.function_code in ADMIN_FUNCTION_ORDER
                else len(ADMIN_FUNCTION_ORDER),
                pc.position_name_ru or pc.position_code,
            )
        )
        for pc in admin_positions:
            parsed = _parsed_for_position(by_pos, pc.position_code)
            if not parsed or not (parsed.hard_skills or parsed.soft_skills):
                continue
            pos_name = (pc.position_name_ru or pc.position_code).strip()
            _append_skills(out, pos_name=pos_name, parsed=parsed)
            seen_positions.add(pc.position_code)
    finally:
        db.close()
    return out


def _style_sheet(ws) -> None:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, ws.max_column + 1):
        max_len = len(str(ws.cell(row=1, column=col).value or ""))
        for row in range(2, min(ws.max_row + 1, 800)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 72)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_xlsx(rows: list[tuple[str, str, str]], dest: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Навыки_стационар"
    ws.append(list(HEADERS))
    for row in rows:
        ws.append(list(row))
    _style_sheet(ws)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def main() -> None:
    rows = collect_rows()
    if not rows:
        raise SystemExit(
            "Нет данных: проверьте docs/regulations/gdrive_hosp, gdrive_default и каталог hosp."
        )
    build_xlsx(rows, OUT_PATH)
    positions = len({r[0] for r in rows})
    print(f"Written {OUT_PATH}")
    print(f"  должностей: {positions}, строк навыков: {len(rows)}")


if __name__ == "__main__":
    main()
